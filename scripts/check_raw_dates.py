import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

fp = r'C:\Users\asus\Downloads\Report AR Finance Test (3).xlsx'
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['Daily']

# Print header
hdr = [c.value for c in ws[1]]
print('Headers:', hdr[:6])

# Find rows with date 5/1 and 5/2
print('\nRows with raw date == 2026-05-01 or 2026-05-02:')
for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
    if row[0] is None: break
    d = row[0]
    if hasattr(d, 'strftime'):
        if d.strftime('%Y-%m-%d') in ('2026-05-01','2026-05-02','2026-04-30'):
            print(f'  type={type(d).__name__} val={d!r} iso={d.isoformat()} bill={row[3]}')

# Also check Excel serial number of one date
print('\nCell format info for A2-A5:')
for r in range(2,6):
    c = ws.cell(row=r, column=1)
    print(f'  A{r}: value={c.value!r}, number_format={c.number_format!r}, type={type(c.value).__name__}')
